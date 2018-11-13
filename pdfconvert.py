import os
import sys
import pdftables_api
from PyPDF2 import PdfFileWriter, PdfFileReader

import xlrd

from openpyxl import Workbook

if len(sys.argv) < 3:
    command = os.path.basename(__file__)
    sys.exit('Usage: {} pdf-file page-number, ...'.format(command))

pdf_input_file = sys.argv[1];
pages_args = ",".join(sys.argv[2:]).replace(" ","")
pages_required = [int(p) for p in filter(None, pages_args.split(","))]

print("Converting pages(start, end page number): {}".format(str(pages_required)[1:-1]))

excel_output_file = pdf_input_file[0: len(pdf_input_file)-4] + '.xlsx'

pages_out_of_range = []
pdf_file_reader = PdfFileReader(open(pdf_input_file, 'rb'))
pdf_file_pages = pdf_file_reader.getNumPages()

for page_number in pages_required:
    if page_number < 1 or page_number > pdf_file_pages:
        pages_out_of_range.append(page_number)

if len(pages_out_of_range) > 0:
    pages_str = str(pages_out_of_range)[1:-1]
    sys.exit('Error: page numbers out of range: {}'.format(pages_str))

pdf_writer_selected_pages = PdfFileWriter()

stp=pages_required[0]
endp=pdf_file_pages

if len(pages_required)==2:
    endp=pages_required[1]


#
# for page_number in pages_required:
#     page = pdf_file_reader.getPage(page_number-1)
#     pdf_writer_selected_pages.addPage(page)

for page_number in range(stp, endp):
    page=pdf_file_reader.getPage(page_number-1)
    pdf_writer_selected_pages._addPage(page)

pdf_file_selected_pages = pdf_input_file + '.tmp'

with open(pdf_file_selected_pages, 'wb') as f:
   pdf_writer_selected_pages.write(f)

c = pdftables_api.Client("idp25x9arr5m")    #This is PDFTable account API
c.xlsx(pdf_file_selected_pages, excel_output_file)
print("Complete")
os.remove(pdf_file_selected_pages)

wbr=xlrd.open_workbook(excel_output_file)
wb=Workbook()

shw = wb.active
shw.title = 'Sheet'
res=["", "", ""]
cnt=0

cell=shw.cell(row=1, column=1)
cell.value="Song Title"

cell=shw.cell(row=2, column=2)
cell.value="Publisher"

cell=shw.cell(row=3, column=3)
cell.value="Writer"

for shr in wbr.sheet_names():
    for i in range(35):
        cnt=cnt+1
        for j in range(3):
            if len(wbr.sheet_by_name(shr).cell_value(i+2, j))>0:
                res[j]=wbr.sheet_by_name(shr).cell_value(i+2, j)

            cell=shw.cell(row=cnt+1, column=j+1)
            cell.value=res[j]

wb.save("res.xlsx")



